import pandas as pd

import tkinter as tk

from tkinter import messagebox

from tkinterdnd2 import DND_FILES, TkinterDnD

import os

from openpyxl import load_workbook

from openpyxl.styles import PatternFill, Font



def run_feature_1(source_path, ref_path):

    # 1. 파일 읽기

    source_all = pd.read_excel(source_path, sheet_name=None)

    ref_all = pd.read_excel(ref_path, sheet_name=None)

    

    source_sheets = list(source_all.keys())

    ref_sheets = list(ref_all.keys())



    # [비교 기준] 참고파일: Sub Item / Entry ID 결합 (첫 번째 시트)

    df_ref = ref_all[ref_sheets[0]]

    df_ref['KEY'] = df_ref["Sub Item"].astype(str) + "/" + df_ref["Entry ID"].astype(str)



    # 업데이트가 필요한 시트 번호들 (2번째=index 1, 3번째=index 2)

    target_sheet_indices = [1, 2]

    updated_cells = {} # 변경된 셀의 위치를 기억 (시트명: [(행, 열), ...])



    for idx in target_sheet_indices:

        if len(source_sheets) > idx:

            sheet_name = source_sheets[idx]

            df = source_all[sheet_name]

            

            # Mandatory 열 위치 확인

            if 'Mandatory' not in df.columns: continue

            mnd_col_idx = df.columns.get_loc('Mandatory')

            

            updated_cells[sheet_name] = []



            for i, row in df.iterrows():

                old_val = str(row.get('Mandatory', ''))

                new_val = old_val

                

                # 로직 적용

                if idx == 1 and str(row.get('LAYOUT', '')).upper() == "SYSDEFINED":

                    new_val = "N"

                else:

                    # [비교 기준] 원본: DOMAIN / VARIABLE_ID 결합

                    s_key = f"{row.get('DOMAIN', '')}/{row.get('VARIABLE_ID', '')}"

                    match = df_ref[df_ref['KEY'] == s_key]

                    

                    if not match.empty:

                        f_val = str(match.iloc[0].get('Sub Item', ''))

                        o_val = str(match.iloc[0].get('Default Missing Query', ''))

                        

                        if idx == 1: # 2번째 시트 로직

                            new_val = "Y/N" if "." in f_val else ("Y" if o_val == "Yes" else "N" if o_val == "No" else o_val)

                        else: # 3번째 시트 로직 (CAT)

                            new_val = "Y" if o_val == "Yes" else "N" if o_val == "No" else o_val



                # 값이 달라진 경우 기록

                if str(new_val) != old_val:

                    df.at[i, 'Mandatory'] = new_val

                    # openpyxl은 1부터 시작, 헤더 포함이므로 i+2

                    updated_cells[sheet_name].append((i + 2, mnd_col_idx + 1))

            

            source_all[sheet_name] = df



    # 2. 기본 저장 (pandas)

    save_p = os.path.join(os.path.dirname(source_path), "업데이트_결과.xlsx")

    with pd.ExcelWriter(save_p, engine='openpyxl') as writer:

        for name, df in source_all.items():

            if 'KEY' in df.columns: df = df.drop(columns=['KEY'])

            df.to_excel(writer, sheet_name=name, index=False)



    # 3. 스타일 적용 (openpyxl)

    wb = load_workbook(save_p)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    red_font = Font(color="FF0000", bold=True)



    for sheet_name, cells in updated_cells.items():

        ws = wb[sheet_name]

        for r, c in cells:

            cell = ws.cell(row=r, column=c)

            cell.fill = yellow_fill

            cell.font = red_font

    

    wb.save(save_p)

    return save_p



# --- UI 부분은 기존과 동일하되 실행 함수만 연결 ---

def process():

    src_p = label2.cget("text").strip('{}')

    ref_p = label3.cget("text").strip('{}')

    if not os.path.exists(src_p) or not os.path.exists(ref_p):

        messagebox.showwarning("알림", "파일을 모두 넣어주세요!")

        return

    try:

        final_path = run_feature_1(src_p, ref_p)

        messagebox.showinfo("성공", f"업데이트 완료 및 스타일 적용!\n파일 위치:\n{final_path}")

    except Exception as e:

        messagebox.showerror("오류", str(e))



# UI 설정 (생략 - 기존 드래그 앤 드롭 코드 사용)

root = TkinterDnD.Tk()

# ... (이전 UI 코드와 동일) ...

root.mainloop()
